import openpyxl

#Carregar planilha
empregados = openpyxl.load_workbook('automacao_excel_1.xlsx')

#Carregar página
empregados_sheets = empregados['Planilha1']



#Ler as linhas
for linha in empregados_sheets.iter_rows(min_col=2, min_row=2):
    #[Empregados, Valor1, Valor2, Valor final, Valor final Correto]
    #Declaração dos dois valores e valor final
    valor1= linha[0].value #Lê a primeira linha
    valor2= linha[1].value #L~e a segunda linha
    valorfinalcorreto = valor1+valor2
    valorfinal = linha[2].value
    #Verificação dos valores, se correto OK, se incorreto, ajuste
    if valorfinal == valorfinalcorreto:
        empregados_sheets.cell(row=2, column = 5).value = 'OK'
    else:
        empregados_sheets.cell(row=2, column = 5).value = valorfinalcorreto
    
for linha in empregados_sheets.iter_rows(max_col=5, min_row=2):
    linha_valorfinal = linha[4].value
    print(linha_valorfinal)

# Salvar projeto
    empregados.save('automacao_excel_1.xlsx')